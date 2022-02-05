from django.shortcuts import render
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django import forms
from . forms import forma
import pyautogui
import time
import pyperclip                # copy paste allowing library
from openpyxl import Workbook, load_workbook # excel library
from django.views.generic import TemplateView
from . models import groups

BGBLACK = '\u001b[40m'
BGGREEN = '\u001b[42m'
BGRED = '\u001b[41m'
CEND = '\033[0m'
# labukas

excel_file = load_workbook('facebook_groups.xlsx')
excel_sheet = excel_file['didieji']

####################################################################################################################################################################

def test(request):
    # first, we import models into this view.
    # from . models import <model name>
    # then, we create a variable that stores a function from db? Like so.
    all_groups = groups.objects.all().count()
    veganai_groups = groups.objects.filter(group_category='veganai').count()
    dovanos_groups = groups.objects.filter(group_category='dovanos').count()
    
    # then, we put that variable into context variable, which then...
    context = {
        'all_groups': all_groups,
        'veganai_groups': veganai_groups,
        'dovanos_groups': dovanos_groups,
    }
    
    items = groups.objects.all()
    for item in items.values('group_name', 'group_link', 'group_category'):
        name = item['group_name']
        link = item['group_link']
        category = item['group_category']
        print(category)

    # is returned at the end. Context = context is the key.
    return render(request, 'main/test.html', context=context)
    # when that is done, I can then go to html templateview and do {{ context variable }}
    # and it prints out on the web! boom.

####################################################################################################################################################################
    
def homepage(request):
    my_form = forma()
    if request.method == "POST":
        my_form = forma(request.POST)
        if my_form.is_valid():
            '''
            Just getting some values for later use
            '''
            form_link = (my_form.cleaned_data['Link'])
            form_text = (my_form.cleaned_data['Text'])
            form_text2 = (my_form.cleaned_data['Text2'])
            form_number = (my_form.cleaned_data['Number'])
            form_category = (my_form.cleaned_data['Category'])
            print("\n")
            print("Dalykai kuriuos irasei yra:")
            print("LINK " + "= " + str( form_link))
            print("TEXT " + "= " + str( form_text))
            print("TEXT2 " + "= " + str( form_text2))
            print("Number " + "= " + str( form_number))
            print("Category " + "= " + str( form_category))
            print("\n")
            print("uz 5 sec procedura prasides")
            print("\n")
            time.sleep(5)

            print("3 Seconds to prepare the browser")
            for i in range(4):
                time.sleep(1)
                print("Prepare browser" + " " + str(i) + "/3")
            time.sleep(1)
            print("Opening browser")
            time.sleep(1)
            print("\n")
            pyautogui.keyDown('ctrl')
            pyautogui.keyDown('t')
            pyautogui.keyUp('t')
            pyautogui.keyUp('ctrl')

            count = 0
            scriptoPradzia = time.time()
            
            for row in excel_sheet.iter_rows(max_row=int(form_number)):
                postoPradzia = time.time()
                group_url = row[1].value  # fetch group id from excel
                group_name = row[0].value # fetch group name from excel
                link = 'https://facebook.com/groups/'+str(group_url) # pro move
                # type group name
                time.sleep(3)
                pyperclip.copy(link)          # copies 'facebook.com/groups' and adds the group ID from excel sheet at the end of this url
                pyautogui.hotkey('ctrl', 'v') # pastes the FULL url
                pyautogui.typewrite('\n')     # press ENTER
                print("Atsidariau" + " " + BGBLACK + str(group_name) + CEND) # prints out in the console the name of the group it has opened
                print("\n")
                
                # let the browser window load
                print("9 seconds to let the browser window load")
                for i in range(10):
                    time.sleep(1)
                    print("Browser window load" + " " + str(i) + "/9")


                try:
                    x, y = pyautogui.locateCenterOnScreen("/home/arvydas/Dropbox/projects/facebook_automated_groups/resources/cpp.png")
                    print("The image 'create_public_post.png' was found.")
                    pyautogui.click(x,y)
                except TypeError:
                    print("Could not locate the image - Create a public post...")
                    a, b = pyautogui.locateCenterOnScreen("/home/arvydas/Dropbox/projects/facebook_automated_groups/resources/ws.png")
                    print("The image 'write something' was found")
                    pyautogui.click(a,b)


                time.sleep(2)
                pyperclip.copy(form_link)
                time.sleep(1)
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.hotkey('ctrl','a') 
                pyautogui.press('backspace') 
                pyperclip.copy(form_text)
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.press('enter')
                pyperclip.copy(form_text2)
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.press('enter')

                time.sleep(1)
                pyautogui.click(1664, 525) 
                time.sleep(1)
                pyautogui.click(1448, 950) 
                                
                # some time to prepare the browser
                print("2 Seconds to prepare the browser")
                for i in range(2):
                    time.sleep(1)
                    print("Posting" + " " + str(i) + "/2")
                    time.sleep(1)
                pyautogui.write(['f6'])     # mark search field, prepare for new link input
                print(BGGREEN + "PAPOSTINTA"+ CEND+ " i " + " " + BGBLACK + str(group_name)+"."+" " + BGRED + "Uztruko {0} sekundes" .format(time.time() - postoPradzia) + CEND)
                print("--------------------------------------------------------------------------------------")
                print("\n")       # new line
                count +=1 # variable will increment every loop iteration
                
            '''
            How long in TOTAL it took for the script to run and how many groups it posted to.
            '''
            print("Papostinau i" + " " + str(count) + " " + "grupes.") # how many groups I have posted to 
            print("Is viso uztruko {0} sekundes" .format(time.time() - scriptoPradzia)) # how long it took for the script to run
   
    context ={ 
        "form": my_form
    }

    return render(request, 'main/home.html', context) # How to render this page
